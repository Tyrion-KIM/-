# tests/test_kpi_dashboard.py
# -*- coding: utf-8 -*-
import pytest
from openpyxl import load_workbook
from scripts.build_kpi_template import build_workbook, OUT
from scripts.build_kpi_dashboard import (MonthSheetError, read_month,
                                         compute_month, render)

@pytest.fixture(scope="module")
def sample_xlsx(tmp_path_factory):
    """模板 → 打分-2026-09 填样例（含 K02=NA 降级场景）→ 任务块打分。"""
    p = tmp_path_factory.mktemp("kpi") / "sample.xlsx"
    wb = build_workbook()
    ws = wb["打分表模板"]; ws.title = "打分-2026-09"; ws["B1"] = "2026-09"
    raw = {  # 指标ID: (i, j, k, available)  数值与 test_kpi_model 对齐
        "K01": (100, 2000, None, True),   # 成本=20欧→75
        "K02": (None, None, None, False), # NA 降级
        "K03": (0.04, None, None, True),  # 上限0.05→100
        "K04": (8, None, None, True),     # 上限7→87.5
        "K05": (0.97, None, None, True),  # 下限0.95→100
        "K06": (0, 0.92, None, True),     # 请款→100
        "K07": (1.0, None, None, True), "K08": (18, 20, None, True),
        "K09": (92, 100, None, True), "K10": (1, 0.85, None, True),
        "K11": (100, 1500, None, True), "K12": (90, 100, None, True),
        "K13": (0.06, None, None, True), "K14": (10, None, None, True),
        "K15": (0.03, 120, 5, True), "K16": (0, 0.9, None, True),
        "K17": (0, 0.95, None, True),
        "K18": (200, 3000, None, True), "K19": (85, 100, None, True),
        "K20": (0.018, None, None, True), "K21": (0.05, 100, 4.5, True),
        "K22": (2, 0.7, None, True),
        "K23": (45, 50, None, True), "K24": (20, 25, None, True),
        "K25": (0, 1.0, None, True),
        "K26": (34.0, None, None, True), "K27": (14.5, 15.2, 5.8, True),
        "K28": (None, None, None, True), "K29": (None, None, None, True),
    }
    from scripts.kpi_model import INDICATORS
    for n, ind in enumerate(INDICATORS):
        r = 4 + n
        i, j, k, ok = raw[ind.id]
        ws.cell(row=r, column=8, value="是" if ok else "NA")
        for col, v in ((9, i), (10, j), (11, k)):
            if v is not None:
                ws.cell(row=r, column=col, value=v)
    tsk = wb["任务块"]
    for r in range(4, 11):          # 7条预埋任务全部打100
        tsk.cell(row=r, column=5, value=100)
    wb.save(p)
    return p

def test_read_compute(sample_xlsx):
    data = read_month(sample_xlsx, "2026-09")
    assert data["rows"]["K01"]["i"] == 100 and data["rows"]["K02"]["available"] is False
    assert data["tasks"]["吴佳钒"] == 100
    res = compute_month(data)
    # 吴佳钒：手算（K01=75×27 + K03=100×5 + K04=87.5×5 + K05=100×5 + K06=100×10 + 任务100×30）
    #        /(27+5+5+5+10+30) = (2025+500+437.5+500+1000+3000)/82 = 7462.5/82
    assert res["totals"]["吴佳钒"] == pytest.approx(91.01, abs=0.05)
    assert res["coverage"]["吴佳钒"] == pytest.approx(52 / 70, abs=0.01)
    assert res["team_cost"] == pytest.approx(min(100, 100 * 36 / 34.0), abs=0.01)  # ≈105.9→100
    assert res["lines"]["吴佳钒"]["DDP头程"] is not None

def test_render_html(sample_xlsx):
    data = read_month(sample_xlsx, "2026-09")
    html = render(compute_month(data))
    for kw in ["2026-09", "吴佳钒", "91.0", "红绿灯", "监控区", "任务块公示",
               "其他项（机动支援）不设卡", "截关装船"]:
        assert kw in html
    assert html.count('class="card light-green"') + \
           html.count('class="card light-amber"') + \
           html.count('class="card light-red"') >= 5

def test_cli_end_to_end(sample_xlsx, tmp_path):
    import subprocess, sys
    out = tmp_path / "d.html"
    subprocess.run([sys.executable, "-X", "utf8",
                    "scripts/build_kpi_dashboard.py", "--month", "2026-09",
                    "--xlsx", str(sample_xlsx), "--out", str(out.parent)],
                   check=True, cwd=str(OUT.parent))
    assert (tmp_path / "kpi_dashboard_2026-09.html").exists()

# --- FUTURE-WORK ②：read_month 校验（月份串 / sheet 存在 / B1 一致 / 环比不吞错） ---

@pytest.mark.parametrize("bad", ["2026-9", "202609", "2026-13"])
def test_read_month_invalid_format(sample_xlsx, bad):
    """月份串不是 YYYY-MM → ValueError，且不 load 工作簿。"""
    with pytest.raises(ValueError, match=r"YYYY-MM"):
        read_month(sample_xlsx, bad)

def test_read_month_missing_sheet(sample_xlsx):
    """打分-YYYY-MM sheet 不存在 → MonthSheetError，消息列出现有打分表。"""
    with pytest.raises(MonthSheetError, match=r"打分-2026-10"):
        read_month(sample_xlsx, "2026-10")

@pytest.mark.parametrize("b1", ["2026-08", None, ""])
def test_read_month_b1_mismatch(tmp_path, b1):
    """B1 绩效月份 ≠ 目标月份（含空）→ ValueError，抓「复制上月没改 B1」。"""
    p = tmp_path / "b1.xlsx"
    wb = build_workbook()
    ws = wb["打分表模板"]; ws.title = "打分-2026-09"; ws["B1"] = b1
    wb.save(p)
    with pytest.raises(ValueError, match="B1"):
        read_month(p, "2026-09")

def test_cli_prev_sheet_b1_mismatch_not_swallowed(tmp_path):
    """上月打分表存在但 B1 错 → 环比不能静默吞掉，CLI 必须失败并报 B1。"""
    import subprocess, sys
    p = tmp_path / "bad_prev.xlsx"
    wb = build_workbook()
    for nm, b1 in (("打分-2026-08", "2026-09"), ("打分-2026-09", "2026-09")):
        ws = wb.copy_worksheet(wb["打分表模板"])
        ws.title = nm; ws["B1"] = b1
    wb.remove(wb["打分表模板"])
    wb.save(p)
    out = tmp_path / "d2"
    r = subprocess.run([sys.executable, "-X", "utf8",
                        "scripts/build_kpi_dashboard.py", "--month", "2026-09",
                        "--xlsx", str(p), "--out", str(out)],
                       capture_output=True, text=True, cwd=str(OUT.parent))
    assert r.returncode != 0
    assert "B1" in r.stderr

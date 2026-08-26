# scripts/build_kpi_template.py
# -*- coding: utf-8 -*-
"""生成 物流团队绩效V1.xlsx（无宏、纯公式、WPS兼容）。用法：python -X utf8 scripts/build_kpi_template.py"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))  # 直跑兼容
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from kpi_model import INDICATORS, PEOPLE, SPECIALISTS

SHEET_GUIDE = "说明"; SHEET_WEIGHTS = "人员与权重"
SHEET_STD = "指标标准"; SHEET_SCORE = "打分表模板"; SHEET_TASK = "任务块"
OUT = Path(__file__).resolve().parent.parent / "物流团队绩效V1.xlsx"

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)

def _hdr(ws, row, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")

def _guide(wb):
    ws = wb.create_sheet(SHEET_GUIDE)
    ws["A1"] = "物流团队绩效 V1 使用说明"; ws["A1"].font = TITLE_FONT
    rows = [
        ("一、月度流程", ""),
        ("1", "月初：复制[打分表模板]并重命名为 打分-YYYY-MM（如 打分-2026-09）"),
        ("2", "月初：在[任务块]登记每人3-5条当月任务+验收物，月份列填当月（如2026-09）"),
        ("3", "月中：日常积累；数据不可得的指标，[数据可用]列选 NA"),
        ("4", "月末：填打分表原始数据；任务块打分（0/50/100）"),
        ("5", "出分：python -X utf8 scripts/build_kpi_dashboard.py --month YYYY-MM"),
        ("二、得分公式", ""),
        ("成本类", "min(100, 100×目标÷实际单台)；单台=运费B÷数量A"),
        ("达标率类", "达标单数÷总单数×100"),
        ("上限类（缺货/退件/库龄等）", "达标100；未达标 100×目标÷实际"),
        ("下限类（准确率）", "达标100；未达标 100×实际÷目标"),
        ("请款类", "时间50%+金额50%；10号前=100，每迟1个工作日-20（下限0）；金额 min(100,占比÷90%×100)"),
        ("三、降级规则（NA）", ""),
        ("规则", "指标数据不可得→[数据可用]选NA→该指标剔除，个人总分按可用权重归一化；没数据不编分数"),
        ("覆盖率", "打分表总分区显示每人 数据覆盖率=可用量化权重÷70；连续两月<60%的线，数据管道建设进该人下月任务块"),
        ("四、可控性原则", ""),
        ("原则", "计分指标必须本人可控；不可控但重要的（FOB截关装船准时率-货代控制）在打分表监控区只显示不计分"),
        ("五、看板红绿灯", ""),
        ("阈值", "绿≥85 / 黄70-84.9 / 红<70 / NA灰；其他项（机动支援）不设卡，只任务块月度考核"),
        ("六、版本", ""),
        ("V1.1", "2026-08-26 定稿：卡结构 6→5（雨洁卡撤销，K23/K24/K25 并入郑舒漫并重新配权）；权重固定，2027-02 复盘"),
    ]
    for r, (a, b) in enumerate(rows, 3):
        ws.cell(row=r, column=1, value=a).font = Font(bold=True)
        ws.cell(row=r, column=2, value=b)
    ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 90

def _weights(wb):
    ws = wb.create_sheet(SHEET_WEIGHTS)
    ws["A1"] = "团队分工与权重（固定，2027-02复盘调整）"; ws["A1"].font = TITLE_FONT
    _hdr(ws, 3, ["姓名", "岗位", "分工", "量化块明细（70%）", "任务块（30%）"])
    team = [
        ("金炜铮", "物流中心主管", "物流中心全范围", "团队单台成本30 · 三段结构20 · 团队请款10 · 专员均分10", "管理任务"),
        ("郑舒漫", "物流专员", "FOB · 特殊发运 · 进口 · 单据（接雨洁）", "FOB(单证7.5+订舱5) · FCA执行12.5 · 请款10 · 特殊发运20 · 进口时效10 · 账单协同5", "30"),
        ("吴佳钒", "物流专员", "头程DDP", "DDP成本27+时效18 · 始发计划10 · 在途库存5 · 请款10", "30"),
        ("黄婷", "物流专员", "目的国大货中转、B端运输", "B端成本21+时效14 · 目的国计划10 · B端库存10 · 请款15", "30"),
        ("吴定佳", "物流专员", "一件代发、C端履约、售后退货", "一件代发成本24+时效16 · 退件率10 · C端库存10 · 请款10", "30"),
    ]
    for r, row in enumerate(team, 4):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    for col, w in zip("ABCDE", (10, 14, 30, 52, 12)):
        ws.column_dimensions[col].width = w

def _std(wb):
    ws = wb.create_sheet(SHEET_STD)
    ws["A1"] = "指标标准字典（唯一口径来源）"; ws["A1"].font = TITLE_FONT
    _hdr(ws, 3, ["指标ID", "人员", "业务线", "指标", "公式类型", "目标", "单位", "权重",
                 "原始字段说明", "数据源", "取值口径"])
    for r, ind in enumerate(INDICATORS, 4):
        vals = [ind.id, ind.person, ind.line, ind.name, ind.ftype, ind.target,
                ind.unit, ind.weight, " / ".join(ind.raw), ind.source, ind.scope]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=v)
    ws.column_dimensions["I"].width = 42; ws.column_dimensions["J"].width = 22
    ws.column_dimensions["K"].width = 34
    ws.freeze_panes = "A4"

SCORE_FIRST_ROW = 4
# 行映射（与 kpi_model.INDICATORS 顺序一致）
ROW_OF = {ind.id: SCORE_FIRST_ROW + n for n, ind in enumerate(INDICATORS)}
PAY_ROWS = ["K06", "K10", "K16", "K17", "K22", "K25"]          # 请款指标
PERSON_RANGE = {  # 每人指标行区间（含端点）
    "吴佳钒": (4, 9), "郑舒漫": (10, 16), "黄婷": (17, 23),
    "吴定佳": (24, 28), "金炜铮": (29, 32),
}
TOTAL_FIRST_ROW = 39   # 总分区人员首行（39-43，5 人，顺序=PEOPLE）

def _m_formula(ind, r):
    t = ind.ftype
    if t == "成本-计算":
        f = f'IF($H{r}<>"是","",IF(OR($I{r}=0,$I{r}=""),"",MIN(100,100*$F{r}/($J{r}/$I{r}))))'
    elif t == "成本-直填":
        f = f'IF($H{r}<>"是","",IF($I{r}="","",MIN(100,100*$F{r}/$I{r})))'
    elif t == "达标率":
        f = f'IF($H{r}<>"是","",IF(OR($J{r}=0,$J{r}=""),"",MIN(100,$I{r}/$J{r}*100)))'
    elif t == "上限":
        f = f'IF($H{r}<>"是","",IF($I{r}="","",IF($I{r}<=$F{r},100,100*$F{r}/$I{r})))'
    elif t == "下限":
        f = f'IF($H{r}<>"是","",IF($I{r}="","",IF($I{r}>=$F{r},100,100*$I{r}/$F{r})))'
    elif t == "请款":
        f = (f'IF($H{r}<>"是","",IF(OR($I{r}="",$J{r}=""),"",'
             f'0.5*MAX(0,100-20*$I{r})+0.5*MIN(100,$J{r}/0.9*100)))')
    elif t == "库存综合":
        f = (f'IF($H{r}<>"是","",IF(OR($I{r}="",$J{r}="",$K{r}=""),"",AVERAGE(IF($I{r}<=0.03,100,100*0.03/$I{r}),'
             f'IF($J{r}<=120,100,100*120/$J{r}),IF($K{r}<=5,100,100*5/$K{r}))))')
    elif t == "三段综合":
        f = (f'IF($H{r}<>"是","",IFERROR(AVERAGE(MIN(100,100*15/$I{r}),'
             f'MIN(100,100*15/$J{r}),MIN(100,100*6/$K{r})),""))')
    elif t == "引用-请款均":
        refs = ",".join(f"$M${ROW_OF[k]}" for k in PAY_ROWS)
        f = f'IF($H{r}<>"是","",IFERROR(AVERAGE({refs}),""))'
    elif t == "引用-专员均分":
        f = f'IF($H{r}<>"是","",IFERROR(AVERAGE($F$39:$F$42),""))'
    else:
        raise ValueError(t)
    return f'=IFERROR({f},"")'

def _score(wb):
    ws = wb.create_sheet(SHEET_SCORE)
    ws["A1"] = "绩效月份："; ws["B1"] = "2026-09"
    ws["C1"] = "复制本表命名为 打分-YYYY-MM；数据不可得选NA；原始数据见[指标标准]字段说明"
    _hdr(ws, 3, ["指标ID", "人员", "业务线", "指标", "公式类型", "目标值", "单位", "数据可用",
                 "原始1", "原始2", "原始3", "原始4", "指标得分", "全局权重", "有效权重", "加权分"])
    for n, ind in enumerate(INDICATORS):
        r = SCORE_FIRST_ROW + n
        for col, v in [(1, ind.id), (2, ind.person), (3, ind.line), (4, ind.name),
                       (5, ind.ftype), (6, ind.target), (7, ind.unit), (8, "是"),
                       (13, _m_formula(ind, r)), (14, ind.weight)]:
            ws.cell(row=r, column=col, value=v)
        ws.cell(row=r, column=15, value=f'=IF(AND($H{r}="是",ISNUMBER($M{r})),$N{r},0)')
        ws.cell(row=r, column=16, value=f'=IF(ISNUMBER($M{r}),$M{r}*$O{r},0)')
        if ind.unit == "%":
            ws.cell(row=r, column=6).number_format = "0.0%"
            if ind.ftype != "达标率":
                ws.cell(row=r, column=9).number_format = "0.0%"
        if ind.ftype == "请款":
            ws.cell(row=r, column=10).number_format = "0.0%"
        if ind.ftype == "库存综合":
            ws.cell(row=r, column=9).number_format = "0.0%"
    # 数据校验：H 列 是/NA
    dv = DataValidation(type="list", formula1='"是,NA"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"H{SCORE_FIRST_ROW}:H{SCORE_FIRST_ROW + len(INDICATORS) - 1}")
    # 监控区（不计分）
    ws.cell(row=34, column=1, value="监控区（不计分，只显示）——不可控但重要").font = Font(bold=True)
    ws.cell(row=35, column=1, value="M01"); ws.cell(row=35, column=2, value="郑舒漫")
    ws.cell(row=35, column=3, value="FOB"); ws.cell(row=35, column=4, value="截关装船准时率")
    ws.cell(row=35, column=5, value="达标率"); ws.cell(row=35, column=8, value="是")
    ws.cell(row=35, column=9, value=None); ws.cell(row=35, column=10, value=None)
    ws.cell(row=35, column=13,
            value='=IF($H35<>"是","",IF(OR($J35=0,$J35=""),"",MIN(100,$I35/$J35*100)))')
    ws.cell(row=35, column=14, value="—")
    # 总分区
    ws.cell(row=37, column=1, value="个人总分区（自动计算）").font = TITLE_FONT
    _hdr(ws, 38, ["人员", "量化有效权重和", "量化加权分和", "任务块得分", "数据覆盖率", "总分"])
    for n, p in enumerate(PEOPLE):
        r = TOTAL_FIRST_ROW + n
        lo, hi = PERSON_RANGE[p]
        ws.cell(row=r, column=1, value=p)
        ws.cell(row=r, column=2, value=f"=SUM($O${lo}:$O${hi})")
        ws.cell(row=r, column=3, value=f"=SUM($P${lo}:$P${hi})")
        ws.cell(row=r, column=4, value=(
            f'=IFERROR(AVERAGEIFS(任务块!$E:$E,任务块!$B:$B,$A{r},'
            f'任务块!$A:$A,$B$1),"")'))
        ws.cell(row=r, column=5, value=f"=B{r}/70").number_format = "0%"
        ws.cell(row=r, column=6, value=(
            f'=IFERROR((C{r}+IF(ISNUMBER(D{r}),D{r}*30,0))'
            f'/(B{r}+IF(ISNUMBER(D{r}),30,0)),"")'))
    ws.freeze_panes = "A4"
    for col, w in zip("ABCDEFGHIJKLMNOP", (8, 8, 10, 16, 10, 8, 8, 9, 9, 9, 9, 9, 9, 8, 8, 9)):
        ws.column_dimensions[col].width = w

SEED_TASKS = [
    ("2026-09", "吴佳钒", "DDP头程交接平稳", "当月零错发（头程大表核对）"),
    ("2026-09", "吴佳钒", "易仓货件数据打通", "易仓导出口径文档+首月数据回填"),
    ("2026-09", "郑舒漫", "FOB货代台账上线", "台账含单证+订舱时效字段、近3个月数据"),
    ("2026-09", "黄婷", "B端库存三项指标基线摸底", "各仓长周期/库龄/库内成本基线表"),
    ("2026-09", "吴定佳", "一件代发时效分段标准定稿", "打包/运输分段标准文档，主管签认"),
    ("2026-09", "郑舒漫", "雨洁工作交接验收（9/7）", "交接文档+实单演练，特殊发运/进口/单据可独立跑"),
    ("2026-09", "郑舒漫", "特殊发运时效分段标准", "分段标准表（无现成标准则注明NA）"),
    ("2026-09", "金炜铮", "绩效体系首月跑通+复盘", "首月记分卡+看板产出+修订清单"),
]

def _task(wb):
    ws = wb.create_sheet(SHEET_TASK)
    ws["A1"] = "任务块（每人每月3-5条，验收物必须可判定；0/50/100=未启动/进行中/完成）"
    ws["A1"].font = TITLE_FONT
    _hdr(ws, 3, ["月份", "姓名", "任务内容", "验收物", "得分"])
    for r, row in enumerate(SEED_TASKS, 4):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
        ws.cell(row=r, column=5, value=None)   # 得分月末打
    dv = DataValidation(type="list", formula1='"0,50,100"', allow_blank=True)
    ws.add_data_validation(dv); dv.add("E4:E200")
    ws.freeze_panes = "A4"
    for col, w in zip("ABCDE", (12, 10, 40, 46, 8)):
        ws.column_dimensions[col].width = w

def build_workbook():
    wb = Workbook(); wb.remove(wb.active)
    _guide(wb); _weights(wb); _std(wb)
    _score(wb)
    _task(wb)
    return wb

def main():
    wb = build_workbook()          # Task 3/4 扩展此函数
    wb.save(OUT)
    print(f"saved: {OUT}")

if __name__ == "__main__":
    sys.exit(main())

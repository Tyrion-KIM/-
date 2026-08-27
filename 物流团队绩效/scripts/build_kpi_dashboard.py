# scripts/build_kpi_dashboard.py
# -*- coding: utf-8 -*-
"""月度看板：读打分表原始数据 → Python 重算 → 单文件 HTML。
用法：python -X utf8 scripts/build_kpi_dashboard.py --month 2026-09 [--xlsx 物流团队绩效V1.xlsx] [--out output/]
注意：不读 Excel 公式结果（无缓存值），一律按 kpi_model 重算。"""
import argparse
import re
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))  # 直跑兼容
from openpyxl import load_workbook
from kpi_model import (INDICATORS, PEOPLE, SPECIALISTS, row_scores,
                       personal_total, score_cost)

ROOT = Path(__file__).resolve().parent.parent
FIRST_ROW = 4
GREEN, AMBER, RED, GRAY = "light-green", "light-amber", "light-red", "light-gray"
MONTH_RE = re.compile(r"^\d{4}-(0[1-9]|1[0-2])$")

class MonthSheetError(ValueError):
    """打分-YYYY-MM sheet 不存在。环比读取仅吞此异常 = 上月未建表（首月）。"""

def _prev_month(m):
    y, mm = map(int, m.split("-"))
    return f"{y - 1}-12" if mm == 1 else f"{y}-{mm - 1:02d}"

def read_month(xlsx, month):
    if not MONTH_RE.match(month):
        raise ValueError(f"月份格式应为 YYYY-MM（如 2026-09），收到：{month!r}")
    wb = load_workbook(xlsx, data_only=False)
    ws_name = f"打分-{month}"
    if ws_name not in wb.sheetnames:
        avail = [n for n in wb.sheetnames if n.startswith("打分-")]
        hint = "、".join(avail) if avail else "无（请先复制「打分表模板」并重命名为 打分-YYYY-MM）"
        raise MonthSheetError(f"工作簿里没有 {ws_name} sheet。现有打分表：{hint}")
    ws = wb[ws_name]
    b1 = ws["B1"].value
    if b1 is None or str(b1).strip() != month:
        disp = "空" if b1 is None else repr(b1)
        raise ValueError(
            f"{ws_name} 的 B1（绩效月份）应为 {month!r}，实际为 {disp}；"
            "请检查是否复制了未更新的上月打分表")
    rows = {}
    for n, ind in enumerate(INDICATORS):
        r = FIRST_ROW + n
        rows[ind.id] = {
            "ind": ind,
            "available": ws.cell(row=r, column=8).value == "是",
            "i": ws.cell(row=r, column=9).value,
            "j": ws.cell(row=r, column=10).value,
            "k": ws.cell(row=r, column=11).value,
        }
    tasks = {}
    task_rows = []            # [(月份,姓名,任务,验收物,得分)] 供任务块公示表
    tw = wb["任务块"]
    for row in tw.iter_rows(min_row=4, values_only=True):
        m, name, _t, _a, sc = (list(row) + [None] * 5)[:5]
        if m == month:
            task_rows.append((m, name, _t, _a, sc))
            if name and isinstance(sc, (int, float)):
                tasks.setdefault(name, []).append(float(sc))
    tasks = {k: sum(v) / len(v) for k, v in tasks.items()}
    return {"month": month, "rows": rows, "tasks": tasks, "task_rows": task_rows}

def compute_month(data):
    rows, tasks = data["rows"], data["tasks"]
    ind_score = {}
    for kid, rec in rows.items():
        ind = rec["ind"]
        if ind.ftype in ("引用-请款均",):
            pay = [ind_score.get(k) for k in ("K06", "K10", "K16", "K17", "K22", "K25")]
            vals = [s for s in pay if s is not None]
            ind_score[kid] = sum(vals) / len(vals) if vals and rec["available"] else None
        elif ind.ftype == "引用-专员均分":
            pass  # 汇总后回填
        else:
            ind_score[kid] = row_scores(ind, rec["i"], rec["j"], rec["k"]) \
                if rec["available"] else None
    totals, lines = {}, {}
    for p in PEOPLE:
        mine = [(ind_score.get(r["ind"].id), r["ind"].weight)
                for r in rows.values() if r["ind"].person == p]
        totals[p] = personal_total(mine, tasks.get(p))
        plines = {}
        for r in rows.values():
            if r["ind"].person == p:
                key = r["ind"].line
                plines.setdefault(key, []).append((ind_score.get(r["ind"].id), r["ind"].weight))
        lines[p] = {k: (sum(s * w for s, w in v if s is not None) /
                        sum(w for s, w in v if s is not None))
                    if any(s is not None for s, _ in v) else None
                    for k, v in plines.items()}
    ind_score["K29"] = (sum(totals[p] for p in SPECIALISTS if totals.get(p) is not None) /
                        len([p for p in SPECIALISTS if totals.get(p) is not None])) \
        if any(totals.get(p) is not None for p in SPECIALISTS) else None
    # 回填 K29 进主管总分
    spec_rows = [(ind_score[r["ind"].id], r["ind"].weight) for r in rows.values()
                 if r["ind"].person == "金炜铮"]
    totals["金炜铮"] = personal_total(spec_rows, tasks.get("金炜铮"))
    # 覆盖率在 K29 回填后统计（引用类指标有值即计权重，与 Excel 模板 O 列口径一致）
    coverage = {}
    for p in PEOPLE:
        cov_w = sum(r["ind"].weight for r in rows.values()
                    if r["ind"].person == p and ind_score.get(r["ind"].id) is not None)
        coverage[p] = cov_w / 70
    k26 = rows["K26"]
    team_cost = score_cost(36, k26["i"]) if k26["available"] and k26["i"] else None
    return {"month": data["month"], "totals": totals, "coverage": coverage,
            "lines": lines, "team_cost": team_cost, "team_actual_cost":
            k26["i"] if k26["available"] else None,
            "task_rows": data.get("task_rows", [])}

def _light(s):
    return GRAY if s is None else GREEN if s >= 85 else AMBER if s >= 70 else RED

def _fmt(s):
    return "NA" if s is None else f"{s:.1f}"

def render(cur, prev_totals=None):
    m = cur["month"]
    ranked = [p for p in PEOPLE if cur["totals"].get(p) is not None]
    team_avg = sum(cur["totals"][p] for p in ranked) / len(ranked) if ranked else None
    def mom(p):
        if not prev_totals or cur["totals"].get(p) is None:
            return "首月"
        prev = prev_totals.get(p)
        if prev is None:
            return "—"
        return f"{cur['totals'][p] - prev:+.1f}"
    all_lines = []
    for p in PEOPLE:
        all_lines += list(cur["lines"].get(p, {}))
    line_order = list(dict.fromkeys(all_lines))
    matrix = ""
    for p in PEOPLE:
        cells = "".join(
            f'<td class="{_light(cur["lines"][p].get(l))}">{_fmt(cur["lines"][p].get(l))}</td>'
            for l in line_order)
        matrix += f"<tr><th>{p}</th>{cells}<td>{cur['coverage'][p]:.0%}</td></tr>"
    head = "".join(f"<th>{l}</th>" for l in line_order)
    cards = ""
    for p in PEOPLE:
        t = cur["totals"].get(p)
        cov = cur["coverage"][p]
        tag = ' <span class="tag">主管</span>' if p == "金炜铮" else ""
        cards += (f'<div class="card {_light(t)}">'
                  f'<div class="name">{p}{tag}</div>'
                  f'<div class="big">{_fmt(t)}</div>'
                  f'<div class="sub">覆盖率 {cov:.0%} · 环比 {mom(p)}</div></div>')
    pct = min(100, 36 / cur["team_actual_cost"] * 100) if cur["team_actual_cost"] else 0
    cost_bar = (f'<div class="bar"><div class="fill" style="width:{pct:.0f}%"></div></div>'
                f'<div>当月单台总成本 {_fmt(cur["team_actual_cost"])} 欧 / 目标 36 欧'
                f' · 得分 {_fmt(cur["team_cost"])}</div>') if cur["team_actual_cost"] \
        else "<div>团队单台总成本：NA</div>"
    tasks_rows = "".join(
        f"<tr><td>{m}</td><td>{n}</td><td>{t}</td><td>{a}</td>"
        f"<td>{'' if s is None else s}</td></tr>"
        for m, n, t, a, s in cur.get("task_rows", []))
    return f"""<!DOCTYPE html><html lang="zh"><head><meta charset="utf-8">
<title>物流团队绩效看板 {m}</title><style>
:root{{--bg:#f6f7f9;--card:#fff;--ink:#1a1d21;--mut:#6b7280;--line:#e5e7eb;
--good:#0ca30c;--warn:#fab219;--bad:#d03b3b;--na:#9e9e9e;
--good-bg:#e7f8e7;--warn-bg:#fdf3d4;--bad-bg:#fdecea;--na-bg:#f5f5f5;--hdr:#1f4e79}}
body{{background:var(--bg);color:var(--ink);font:14px/1.6 system-ui,-apple-system,"Segoe UI","Microsoft YaHei","PingFang SC",sans-serif;margin:0;padding:24px}}
.wrap{{max-width:1080px;margin:0 auto}} h1{{font-size:20px;margin:0 0 4px}}
.mut{{color:var(--mut);font-size:12px}} .grid{{display:grid;grid-template-columns:repeat(6,1fr);gap:12px;margin:16px 0}}
.card{{background:var(--card);border:1px solid var(--line);border-radius:10px;padding:14px}}
.name{{font-weight:600}} .big{{font-size:30px;font-weight:700;font-variant-numeric:tabular-nums}} .sub{{font-size:12px;color:var(--mut)}}
.tag{{font-size:11px;background:#eef2f7;border-radius:4px;padding:1px 6px;color:var(--mut);font-weight:400}}
.light-green{{border-left:4px solid var(--good)}} .light-amber{{border-left:4px solid var(--warn)}}
.light-red{{border-left:4px solid var(--bad)}} .light-gray{{border-left:4px solid var(--na)}}
table{{border-collapse:collapse;width:100%;background:var(--card);border:1px solid var(--line);border-radius:10px}}
th,td{{padding:8px 10px;border-bottom:1px solid var(--line);text-align:center;font-size:13px;font-variant-numeric:tabular-nums}}
th{{background:var(--hdr);color:#fff}} td.light-green{{background:var(--good-bg)}} td.light-amber{{background:var(--warn-bg)}}
td.light-red{{background:var(--bad-bg)}} td.light-gray{{background:var(--na-bg);color:var(--na)}}
section{{margin:20px 0}} h2{{font-size:15px;margin:0 0 8px}} .bar{{height:14px;background:var(--line);border-radius:7px;overflow:hidden}}
.fill{{height:100%;background:var(--good)}}
@media (prefers-color-scheme: dark){{
:root:not([data-theme="light"]){{--bg:#0d0d0d;--card:#1a1a19;--ink:#fff;--mut:#c3c2b7;--line:#2c2c2a;
--good-bg:#0d2b0d;--warn-bg:#332608;--bad-bg:#33130f;--na-bg:#262626;--hdr:#123a5e}}
.tag{{background:#2c2c2a;color:#c3c2b7}}}}
</style></head><body><div class="wrap">
<h1>物流团队绩效看板 · {m}</h1>
<div class="mut">绿≥85 · 黄70-84.9 · 红&lt;70 · NA灰（降级归一化）｜量化70+任务30｜其他项（机动支援）不设卡</div>
<section><h2>团队</h2><div class="card">团队总分：{_fmt(team_avg)} 分</div>
<div style="margin-top:8px">{cost_bar}</div></section>
<section><h2>个人总分</h2><div class="grid">{cards}</div></section>
<section><h2>业务线红绿灯矩阵（过程管理）</h2>
<table><tr><th>人员</th>{head}<th>覆盖率</th></tr>{matrix}</table></section>
<section><h2>监控区（不计分）</h2><table><tr><th>指标</th><th>结果</th></tr>
<tr><td>FOB 截关装船准时率（货代控制）</td><td>见打分表 M01</td></tr></table></section>
<section><h2>任务块公示</h2><table><tr><th>月份</th><th>姓名</th><th>任务</th><th>验收物</th><th>得分</th></tr>{tasks_rows}</table></section>
<div class="mut">物流团队绩效V1 · 生成于 python -X utf8 scripts/build_kpi_dashboard.py</div>
</div></body></html>"""

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--month", required=True)
    ap.add_argument("--xlsx", default=str(ROOT / "物流团队绩效V1.xlsx"))
    ap.add_argument("--out", default=str(ROOT / "output"))
    a = ap.parse_args()
    data = read_month(a.xlsx, a.month)
    cur = compute_month(data)
    prev = None
    pm = _prev_month(a.month)
    try:
        prev = compute_month(read_month(a.xlsx, pm))["totals"]
    except MonthSheetError:
        pass  # 上月打分表未建 → 环比显示「首月」
    out = Path(a.out); out.mkdir(parents=True, exist_ok=True)
    f = out / f"kpi_dashboard_{a.month}.html"
    f.write_text(render(cur, prev), encoding="utf-8")
    print(f"saved: {f}")

if __name__ == "__main__":
    sys.exit(main())

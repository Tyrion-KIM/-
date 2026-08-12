"""build_fba_cost_model.py — 美国站 FBA 全链路成本数据提取与组装"""
import json
import os
from datetime import date
from openpyxl import load_workbook


# === 配置 ===
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "AGL 2026年8月8日生效美线日线海运价格.xlsx")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "fba_us_cost_model.json")


# === 工具函数 ===

def safe_str(v):
    return str(v).strip() if v else ""


def safe_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


# === AGL 数据提取 ===

def _extract_sheet(wb, sheet_name: str, shipment_mode: str) -> list:
    """通用 Sheet 提取。FCL 和 LCL 列结构相同（前 12 列），仅费率列不同。"""
    ws = wb[sheet_name]
    routes = []

    is_fcl = (shipment_mode == "FCL")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 12:
            continue
        if not row[1] or safe_str(row[1]) != "US":
            continue

        route = {
            "shipment_mode": shipment_mode,
            "origin_port": safe_str(row[6]),
            "dest_region": safe_str(row[7]),
            "dest_city": safe_str(row[8]),
            "amazon_fc": safe_str(row[9]),
            "fc_type": safe_str(row[10]),
            "speed_mode": safe_str(row[2]),
            "fob": safe_str(row[5]),
            "currency": safe_str(row[3]),
            "product": safe_str(row[4]),
            "fixed_fee": safe_float(row[11]),
        }

        if is_fcl:
            route["rate_20gp"] = safe_float(row[12])
            route["rate_40gp"] = safe_float(row[13])
            route["rate_40hq"] = safe_float(row[14])
        else:
            route["rate_1_5_cbm"] = safe_float(row[12])
            route["rate_5_10_cbm"] = safe_float(row[13])
            route["rate_10_15_cbm"] = safe_float(row[14])
            route["rate_gt15_cbm"] = safe_float(row[15])

        routes.append(route)

    return routes


def extract_agl(excel_path: str) -> dict:
    """从 AGL 海运价卡 Excel 提取所有路由（FCL 整柜 + LCL 散货）"""
    wb = load_workbook(excel_path, data_only=True)

    fcl_routes = _extract_sheet(wb, "FCL价卡", "FCL")
    lcl_routes = _extract_sheet(wb, "LCL价卡", "LCL")
    wb.close()

    all_routes = fcl_routes + lcl_routes

    def _uniq(key):
        return sorted(set(r[key] for r in all_routes if r.get(key)))

    return {
        "description": "AGL海运头程价卡 — 中国→美国 (含FCL整柜+LCL散货)",
        "source_file": os.path.basename(excel_path),
        "valid_from": "2026-08-08",
        "summary": {
            "total_routes": len(all_routes),
            "fcl_routes": len(fcl_routes),
            "lcl_routes": len(lcl_routes),
            "origin_ports": _uniq("origin_port"),
            "dest_regions": _uniq("dest_region"),
            "dest_cities": _uniq("dest_city"),
            "fc_types": _uniq("fc_type"),
            "fob_types": _uniq("fob"),
            "currencies": _uniq("currency"),
            "speed_modes": _uniq("speed_mode"),
            "products": _uniq("product"),
        },
        "routes": all_routes
    }


# === FBA 配送费 ===

def build_fulfillment() -> dict:
    """FBA 配送费 — 美国站 2026 费率"""
    return {
        "description": "FBA配送费 — 美国站 2026（2026/1/15生效）",
        "currency": "USD",
        "source_url": "https://sellercentral.amazon.com/help/hub/reference/external/GMUTB89XM7AATPR3",
        "sources_verified": [
            "https://docs.m2ecloud.com/docs/amazon-fba-fulfillment-fees-in-the-us-explained/",
            "https://www.inventoryhero.ai/blog/fba-fulfillment-fees",
            "https://amzprep.com/holiday-peak-fulfillment-fees/"
        ],
        "fuel_surcharge_pct": 3.5,
        "fuel_surcharge_effective": "2026-04-17",
        "effective_period": {
            "non_peak": "2026-01-15 to 2026-10-14",
            "peak": "2026-10-15 to 2027-01-14"
        },
        "price_bands": {
            "under_10": "售价 < $10 (低价FBA折扣)",
            "10_to_50": "售价 $10-$50",
            "over_50": "售价 > $50"
        },
        "size_tiers": _fulfillment_tiers()
    }


def _fulfillment_tiers() -> list:
    """构建完整的尺寸分段费率数据"""

    small_std = {
        "tier_name": "小号标准 (Small Standard)",
        "tier_name_cn": "小号标准",
        "size_limits": "≤15×12×0.75 in, ≤16 oz",
        "is_apparel": False,
        "weight_breakpoints": [
            {"max_weight_oz": 2,  "non_peak": {"under_10": 2.43, "10_to_50": 3.32, "over_50": 3.58}, "peak": {"10_to_50": 3.51}},
            {"max_weight_oz": 4,  "non_peak": {"under_10": 2.49, "10_to_50": 3.42, "over_50": 3.68}, "peak": {"10_to_50": 3.61}},
            {"max_weight_oz": 6,  "non_peak": {"under_10": 2.56, "10_to_50": 3.45, "over_50": 3.71}, "peak": {"10_to_50": 3.65}},
            {"max_weight_oz": 8,  "non_peak": {"under_10": 2.66, "10_to_50": 3.54, "over_50": 3.80}, "peak": {"10_to_50": 3.74}},
            {"max_weight_oz": 10, "non_peak": {"under_10": 2.77, "10_to_50": 3.68, "over_50": 3.94}, "peak": {"10_to_50": 3.89}},
            {"max_weight_oz": 12, "non_peak": {"under_10": 2.82, "10_to_50": 3.78, "over_50": 4.04}, "peak": {"10_to_50": 3.99}},
            {"max_weight_oz": 14, "non_peak": {"under_10": 2.92, "10_to_50": 3.91, "over_50": 4.17}, "peak": {"10_to_50": 4.13}},
            {"max_weight_oz": 16, "non_peak": {"under_10": 2.95, "10_to_50": 3.96, "over_50": 4.22}, "peak": {"10_to_50": 4.18}}
        ]
    }

    small_std_apparel = {
        "tier_name": "小号标准-服装 (Small Standard - Apparel)",
        "tier_name_cn": "小号标准(服装)",
        "size_limits": "≤15×12×0.75 in, ≤16 oz",
        "is_apparel": True,
        "weight_breakpoints": [
            {"max_weight_oz": 2,  "non_peak": {}, "peak": {"under_10": 2.73, "10_to_50": 3.50, "over_50": 3.50}},
            {"max_weight_oz": 4,  "non_peak": {}, "peak": {"under_10": 2.73, "10_to_50": 3.50, "over_50": 3.50}},
            {"max_weight_oz": 6,  "non_peak": {}, "peak": {"under_10": 2.90, "10_to_50": 3.67, "over_50": 3.67}},
            {"max_weight_oz": 8,  "non_peak": {}, "peak": {"under_10": 2.90, "10_to_50": 3.67, "over_50": 3.67}},
            {"max_weight_oz": 10, "non_peak": {}, "peak": {"under_10": 3.22, "10_to_50": 3.99, "over_50": 3.99}},
            {"max_weight_oz": 12, "non_peak": {}, "peak": {"under_10": 3.22, "10_to_50": 3.99, "over_50": 3.99}},
            {"max_weight_oz": 14, "non_peak": {}, "peak": {"under_10": 3.50, "10_to_50": 4.27, "over_50": 4.27}},
            {"max_weight_oz": 16, "non_peak": {}, "peak": {"under_10": 3.50, "10_to_50": 4.27, "over_50": 4.27}}
        ],
        "_note": "服装非旺季费率待补充"
    }

    large_std = {
        "tier_name": "大号标准 (Large Standard)",
        "tier_name_cn": "大号标准",
        "size_limits": "≤18×14×8 in, ≤20 lb",
        "is_apparel": False,
        "weight_breakpoints": [
            {"max_weight_oz": 4,   "non_peak": {"under_10": 2.91, "10_to_50": 3.73, "over_50": 3.99}, "peak": {"10_to_50": 3.97}},
            {"max_weight_oz": 8,   "non_peak": {"under_10": 3.13, "10_to_50": 3.95, "over_50": 4.21}, "peak": {"10_to_50": 4.21}},
            {"max_weight_oz": 12,  "non_peak": {"under_10": 3.38, "10_to_50": 4.20, "over_50": 4.46}, "peak": {"10_to_50": 4.48}},
            {"max_weight_oz": 16,  "non_peak": {"under_10": 3.78, "10_to_50": 4.60, "over_50": 4.86}, "peak": {"10_to_50": 4.89}},
            {"max_weight_lb": 1.25, "non_peak": {"under_10": 4.22, "10_to_50": 5.04, "over_50": 5.30}, "peak": {"10_to_50": 5.34}},
            {"max_weight_lb": 1.50, "non_peak": {"under_10": 4.60, "10_to_50": 5.42, "over_50": 5.68}, "peak": {"10_to_50": 5.73}},
            {"max_weight_lb": 1.75, "non_peak": {"under_10": 4.75, "10_to_50": 5.57, "over_50": 5.83}, "peak": {"10_to_50": 5.89}},
            {"max_weight_lb": 2.00, "non_peak": {"under_10": 5.00, "10_to_50": 5.82, "over_50": 6.08}, "peak": {"10_to_50": 6.15}},
            {"max_weight_lb": 2.25, "non_peak": {"under_10": 5.10, "10_to_50": 5.92, "over_50": 6.18}, "peak": {"10_to_50": 6.29}},
            {"max_weight_lb": 2.50, "non_peak": {"under_10": 5.28, "10_to_50": 6.10, "over_50": 6.36}, "peak": {"10_to_50": 6.49}},
            {"max_weight_lb": 2.75, "non_peak": {"under_10": 5.44, "10_to_50": 6.26, "over_50": 6.52}},
            {"max_weight_lb": 3.00, "non_peak": {"under_10": 5.85, "10_to_50": 6.67, "over_50": 6.93}},
            {"max_weight_lb": 20.0, "non_peak": {"under_10": "6.15 + 0.08/4oz above 3lb", "10_to_50": "6.97 + 0.08/4oz above 3lb", "over_50": "7.23 + 0.08/4oz above 3lb"}}
        ]
    }

    large_std_apparel = {
        "tier_name": "大号标准-服装 (Large Standard - Apparel)",
        "tier_name_cn": "大号标准(服装)",
        "size_limits": "≤18×14×8 in, ≤20 lb",
        "is_apparel": True,
        "weight_breakpoints": [
            {"max_weight_oz": 4,   "non_peak": {}, "peak": {"under_10": 3.79, "10_to_50": 4.56, "over_50": 4.56}},
            {"max_weight_oz": 8,   "non_peak": {}, "peak": {"under_10": 4.00, "10_to_50": 4.77, "over_50": 4.77}},
            {"max_weight_oz": 12,  "non_peak": {}, "peak": {"under_10": 4.23, "10_to_50": 5.00, "over_50": 5.00}},
            {"max_weight_oz": 16,  "non_peak": {}, "peak": {"under_10": 4.69, "10_to_50": 5.46, "over_50": 5.46}},
            {"max_weight_lb": 1.25, "non_peak": {}, "peak": {"under_10": 5.50, "10_to_50": 6.27, "over_50": 6.27}},
            {"max_weight_lb": 1.50, "non_peak": {}, "peak": {"under_10": 5.50, "10_to_50": 6.27, "over_50": 6.27}},
            {"max_weight_lb": 1.75, "non_peak": {}, "peak": {"under_10": 5.76, "10_to_50": 6.53, "over_50": 6.53}},
            {"max_weight_lb": 2.00, "non_peak": {}, "peak": {"under_10": 5.76, "10_to_50": 6.53, "over_50": 6.53}},
            {"max_weight_lb": 2.25, "non_peak": {}, "peak": {"under_10": 6.27, "10_to_50": 7.04, "over_50": 7.04}},
            {"max_weight_lb": 2.50, "non_peak": {}, "peak": {"under_10": 6.27, "10_to_50": 7.04, "over_50": 7.04}},
            {"max_weight_lb": 2.75, "non_peak": {}, "peak": {"under_10": 6.50, "10_to_50": 7.27, "over_50": 7.27}},
            {"max_weight_lb": 3.00, "non_peak": {}, "peak": {"under_10": 6.50, "10_to_50": 7.27, "over_50": 7.27}},
            {"max_weight_lb": 20.0, "non_peak": {}, "peak": {"under_10": "6.82 + 0.16/0.5lb above 3lb", "10_to_50": "7.59 + 0.16/0.5lb above 3lb", "over_50": "7.59 + 0.16/0.5lb above 3lb"}}
        ],
        "_note": "服装非旺季费率待补充"
    }

    small_bulky = {
        "tier_name": "小号大件 (Small Bulky)",
        "tier_name_cn": "小号大件",
        "size_limits": "≤60 in, L+girth≤130 in, ≤50 lb",
        "is_apparel": False,
        "weight_breakpoints": [
            {"max_weight_lb": 50.0,
             "non_peak": {"under_10": "6.78 + 0.38/lb above 1st lb", "10_to_50": "7.55 + 0.38/lb above 1st lb", "over_50": "7.55 + 0.38/lb above 1st lb"},
             "peak": {"under_10": "9.88 + 0.38/lb above 1st lb", "10_to_50": "10.65 + 0.38/lb above 1st lb", "over_50": "10.65 + 0.38/lb above 1st lb"}}
        ]
    }

    large_bulky = {
        "tier_name": "大号大件 (Large Bulky)",
        "tier_name_cn": "大号大件",
        "size_limits": ">小号大件限制, ≤50 lb",
        "is_apparel": False,
        "weight_breakpoints": [
            {"max_weight_lb": 50.0,
             "non_peak": {"under_10": "8.58 + 0.38/lb above 1st lb", "10_to_50": "9.35 + 0.38/lb above 1st lb", "over_50": "9.35 + 0.38/lb above 1st lb"},
             "peak": {"under_10": "9.88 + 0.38/lb above 1st lb", "10_to_50": "10.65 + 0.38/lb above 1st lb", "over_50": "10.65 + 0.38/lb above 1st lb"}}
        ]
    }

    extra_large = {
        "tier_name": "超大件 (Extra Large)",
        "tier_name_cn": "超大件",
        "size_limits": "超出大件限制，按重量分段",
        "is_apparel": False,
        "sub_tiers": [
            {"weight_range": "0-50 lb",
             "non_peak": {"under_10": "25.56 + 0.38/lb above 1st lb", "10_to_50": "26.33 + 0.38/lb above 1st lb", "over_50": "26.33 + 0.38/lb above 1st lb"},
             "peak": {"under_10": "28.29 + 0.38/lb above 1st lb", "10_to_50": "29.06 + 0.38/lb above 1st lb", "over_50": "29.06 + 0.38/lb above 1st lb"}},
            {"weight_range": "50-70 lb",
             "non_peak": {"under_10": "36.55 + 0.75/lb above 51 lb", "10_to_50": "37.32 + 0.75/lb above 51 lb", "over_50": "37.32 + 0.75/lb above 51 lb"},
             "peak": {"under_10": "42.16 + 0.75/lb above 51 lb", "10_to_50": "42.93 + 0.75/lb above 51 lb", "over_50": "42.93 + 0.75/lb above 51 lb"}},
            {"weight_range": "70-150 lb",
             "non_peak": {"under_10": "50.55 + 0.75/lb above 71 lb", "10_to_50": "51.32 + 0.75/lb above 71 lb", "over_50": "51.32 + 0.75/lb above 71 lb"}},
            {"weight_range": "150+ lb",
             "non_peak": {"under_10": "194.18 + 0.19/lb above 151 lb", "10_to_50": "194.95 + 0.19/lb above 151 lb", "over_50": "194.95 + 0.19/lb above 151 lb"}}
        ]
    }

    return [
        small_std, small_std_apparel,
        large_std, large_std_apparel,
        small_bulky, large_bulky, extra_large
    ]


# === FBA 仓储费 ===

def build_storage() -> dict:
    """FBA 月度仓储费 — 美国站 2026"""
    return {
        "description": "FBA月度仓储费 — 美国站",
        "currency": "USD",
        "unit": "per_cubic_foot_per_month",
        "source_url": "https://sellercentral.amazon.com/help/hub/reference/external/G200612770",
        "sources_verified": [
            "https://skucompass.com/amazon-fba-storage-fees-2026/",
            "https://www.inventoryhero.ai/blog/fba-storage-fees-2026"
        ],
        "cubic_foot_formula": "(L × W × H in inches) ÷ 1728",
        "billing": "每月7-15日扣上月费用，按日均立方英尺计算",
        "rates": [
            {
                "category": "标准件 (Standard)",
                "category_cn": "标准件",
                "non_peak_jan_sep": 0.78,
                "peak_oct_dec": 2.40
            },
            {
                "category": "大件 (Oversize)",
                "category_cn": "大件",
                "non_peak_jan_sep": 0.56,
                "peak_oct_dec": 1.40
            }
        ]
    }


# === 组装 ===

def assemble(agl: dict, fulfillment: dict, storage: dict) -> dict:
    return {
        "meta": {
            "version": "2.0",
            "generated_at": date.today().isoformat(),
            "source_urls": {
                "fba_fulfillment": "https://sellercentral.amazon.com/help/hub/reference/external/GMUTB89XM7AATPR3",
                "fba_fee_changes": "https://sellercentral.amazon.com/help/hub/reference/external/ABBX6GZPA8MSZGW",
                "fba_storage": "https://sellercentral.amazon.com/help/hub/reference/external/G200612770",
                "chinese_summary": "https://gs.amazon.cn/news/news-notices-251016"
            },
            "effective_date": "2026-01-15",
            "agl_effective_date": "2026-08-08",
            "notes": "美国站FBA全链路成本核心项：AGL海运头程(含FCL整柜+LCL散货) + FBA配送费 + FBA月度仓储费。不含长期仓储/移除/退货/低库存等附加费。AGL币种以USD为主(含少量RMB)。"
        },
        "agl_ocean_freight": agl,
        "fba_fulfillment": fulfillment,
        "fba_storage": storage
    }


# === 入口 ===

def main():
    print("Extracting AGL rates...")
    agl = extract_agl(EXCEL_PATH)
    s = agl["summary"]
    print(f"  -> {s['total_routes']} routes (FCL:{s['fcl_routes']} + LCL:{s['lcl_routes']})")
    print(f"  -> Origins: {', '.join(s['origin_ports'])}")
    print(f"  -> Dest regions: {', '.join(s['dest_regions'])}")
    print(f"  -> Dest cities: {', '.join(s['dest_cities'])}")
    print(f"  -> FC types: {', '.join(s['fc_types'])}")
    print(f"  -> FOB types: {', '.join(s['fob_types'])}")
    print(f"  -> Currencies: {', '.join(s['currencies'])}")
    print(f"  -> Speed modes: {', '.join(s['speed_modes'])}")
    print(f"  -> Products: {', '.join(s['products'])}")

    print("Building fulfillment rates...")
    fulfillment = build_fulfillment()

    print("Building storage rates...")
    storage = build_storage()

    print("Assembling...")
    model = assemble(agl, fulfillment, storage)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(model, f, ensure_ascii=False, indent=2)

    print(f"Done -> {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
